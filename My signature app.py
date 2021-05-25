import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter.filedialog import askopenfilename, asksaveasfilename
import time
from datetime import datetime

import os
import subprocess
import sys

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl_image_loader import SheetImageLoader
    if 'PIL' not in sys.modules:
        print ("Pillow shall be installed to run the script")
        subprocess.check_call([sys.executable, "-m", "pip", "install", 'Pillow', '--user'])
except ImportError:
    print ("openpyxl shall be installed to run the script")
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'openpyxl', '--user'])
    print ("openpyxl_image_loader shall be installed to run the script")
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'openpyxl_image_loader', '--user'])
    if 'PIL' in sys.modules:
        print ("Pillow shall be installed to run the script")
        subprocess.check_call([sys.executable, "-m", "pip", "install", 'Pillow', '--user'])
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl_image_loader import SheetImageLoader

try:
    import fitz
except ImportError:
    print ("fitz shall be installed to run the script")
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'fitz', '--user'])
    print ("PyMuPDF shall be installed to run the script")
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'PyMuPDF', '--user'])
    import fitz

try:
    import win32com.client
except ImportError:
    print ("win32com shall be installed to export excel to pdf")
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'pywin32', '--user'])
    import win32com.client

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os.path


class RegDialog(tk.Toplevel):
    # A dialog box to get registration number
    # This is not needed unless the car was speeding
    def __init__(self,master,*args,**kwargs):
        tk.Toplevel.__init__(self,master,*args,**kwargs)
        
        #position window over parent       
        self.geometry("350x100+%d+%d" % (master.winfo_rootx(),master.winfo_rooty()))
        
        self.title("Write your email credentials")
        tk.Label(self,text = "Email",height = 2).grid(row = 0,column = 0)
        tk.Label(self,text = "Password",height = 2).grid(row = 1,column = 0)
        
        self.emailtext = tk.StringVar()
        entry = tk.Entry(self,textvariable = self.emailtext)
        entry.grid(row = 0,column = 1)
        entry.focus_set()

        self.paswordtext = tk.StringVar()
        entry = tk.Entry(self,textvariable = self.paswordtext,show="*")
        entry.grid(row = 1,column = 1)
        entry.focus_set()
        
        tk.Button(self,text = "Send email", command = self.on_ok).grid(row = 2,column = 0)
        tk.Button(self,text = "Cancel", command = self.on_cancel).grid(row = 2,column = 1)

        try:
            conf_file = open(os.path.dirname(sys.argv[0]) + "\\configuration.txt", 'r')
            temp = conf_file.read().splitlines()
            # close the file with the configuration
            conf_file.close()
            self.emailtext.set(temp[0])
            self.paswordtext.set(temp[1])
        except:
            pass
 
        # make dialog modal
        self.transient(master)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW",self.on_cancel)
        # instead of mainloop call wait_window
        self.wait_window()
        
    def on_cancel(self):
        self.action = False
        self.destroy()
    
    def on_ok(self):
        # called when ok button is pressed
        self.action = True
        self.destroy()


def send_email(email_sender, email_password, email_recipient, email_subject, email_message, attachment_location = ''):
    msg = MIMEMultipart()
    msg['From'] = email_sender
    msg['To'] = email_recipient
    msg['Subject'] = email_subject

    msg.attach(MIMEText(email_message, 'plain'))

    ret = True

    if attachment_location != '':
        filename = os.path.basename(attachment_location)
        attachment = open(attachment_location, "rb")
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        "attachment; filename= %s" % filename)
        msg.attach(part)

    try:
        server = smtplib.SMTP('smtp.office365.com', 587)
        server.ehlo()
        server.starttls()
        server.login(email_sender, email_password)
        text = msg.as_string()
        server.sendmail(email_sender, email_recipient, text)
        # print('email sent')
        server.quit()
    except:
        return False
    return ret

def getFolderPath():
    folder_selected = filedialog.askdirectory()
    folderPath.set(folder_selected)

def getSignaturePath():
    signature_selected = filedialog.askopenfilename()
    signaturePath.set(signature_selected)

def getEmailsPath():
    emails_selected = filedialog.askopenfilename()
    emailsPath.set(emails_selected)

def add_signatures():
    folder = folderPath.get()
    signature = signaturePath.get()
    if not folder:
        txt_show.insert(tk.END, "Select a folder first! " + "\n", 'error')
    elif not signature:
        txt_show.insert(tk.END, "Select a signature! " + "\n", 'error')
    else:
        folder = folder.replace('/','\\')
        output_path = folder
        output_path = output_path + '\\output'
        signature_img = signature.replace('/','\\')
        #create the output folder if it's not there
        if os.path.exists(output_path) == False:
            os.mkdir(output_path)
            txt_show.insert(tk.END, "Output folder has been created at this location: " + output_path + "\n", 'OK')

        for file in os.listdir(folder):
            if os.path.splitext(file)[1] == '.pdf':
                
                # read in pdf
                doc = fitz.open(folder + '\\' + file)

                for page in doc:
                    # search for word APPROVED
                    text = "APPROVED"
                    text_instances = page.searchFor(text)

                    # when found the word add the signature under it
                    for inst in text_instances:
                        signature_location = inst
                        signature_location.x0 = signature_location.x0 - 30
                        signature_location.y0 = signature_location.y0 + 30
                        signature_location.x1 = signature_location.x1 + 30
                        signature_location.y1 = signature_location.y1 + 50
                        page.insertImage(inst, filename = signature_img)

                # save it in the output folder
                doc.save(output_path+"\\" + file, garbage=4, deflate=True, clean=True)
                txt_show.insert(tk.END, "Signature added for: " + file + "\n", 'OK')

            if os.path.splitext(file)[1] == '.xlsx':
                wb = load_workbook(folder + '\\' + file)

                sheet = wb.active
                sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
                # Put the sheet in the image loader
                image_loader = SheetImageLoader(sheet)
                xlsx_export = False
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value == "APPROVED":
                            column = openpyxl.utils.cell.get_column_letter(cell.column+1)
                            signature_cell = column + str(cell.row + 3)
                            # if there's not a signature already in the cell then put it
                            if image_loader.image_in(signature_cell) == False:
                                img = openpyxl.drawing.image.Image(signature_img)
                                img.height = img.height * 3 / 4
                                img.width = img.width * 3 / 4
                                img.anchor = signature_cell
                                sheet.add_image(img)
                                txt_show.insert(tk.END, "Signature added for: " + file + "\n", 'OK')
                            else:
                                txt_show.insert(tk.END, "File is already signed: " + file + "\n", 'warning')
                            xlsx_export = True

                wb.save(folder + '\\' + file)

                if xlsx_export == True:
                    o = win32com.client.Dispatch("Excel.Application")
                    o.Visible = False
                    o.DisplayAlerts = False
                    # pathname = os.path.dirname(sys.argv[0]) 
                    wb = o.Workbooks.Open(folder + '\\' + file)


                    # export the excel as pdf in the output folder
                    wb.ActiveSheet.ExportAsFixedFormat(0, output_path + '\\' + os.path.splitext(file)[0] + '.pdf')
                    o.Quit()
                    txt_show.insert(tk.END, "File exported to: " + file + "\n", 'OK')
                else:
                    txt_show.insert(tk.END, "Signature not added to: " + file + "\n", 'warning')

def send_emails_action():
    folder = folderPath.get()
    if not folder:
        txt_show.insert(tk.END, "Select the folder!" + "\n", 'error')
    else:
        folder = folder.replace('/','\\')
        output_path = folder
        output_path = output_path + '\\output'

        emails_path = emailsPath.get()

        if not emails_path:
            txt_show.insert(tk.END, "select the email list!" + "\n", 'error')
        else:
            dlg = RegDialog(window)
            email_sender = dlg.emailtext.get()
            password_sender = dlg.paswordtext.get()

            if dlg.action == True:
                if not email_sender or not password_sender:
                    txt_show.insert(tk.END, "Put the credentials for email!" + "\n", 'error')
                else:
                    try:
                        conf_file = open(os.path.dirname(sys.argv[0]) + "\\configuration.txt", 'r')
                        # close the file with the configuration
                        conf_file.close()
                    except:
                        conf_file = open(os.path.dirname(sys.argv[0]) + "\\configuration.txt", 'w')
                        conf_file.write(email_sender+"\n")
                        conf_file.write(password_sender)
                        # close the file with the configuration
                        conf_file.close()
                    for file in os.listdir(output_path):
                        if os.path.splitext(file)[1] == '.pdf':
                            
                            # read in pdf
                            doc = fitz.open(output_path + '\\' + file)

                            for page in doc:
                                # Open the file with the names and emails
                                emails_file = open(emails_path, 'r')

                                # split the document by lines
                                temp = emails_file.read().splitlines()

                                # go throw all names to find wich one is in the pdf
                                for line in temp:
                                    # split the name and email
                                    name = []
                                    name = line.split("/")

                                    if name:
                                        text_instances = page.searchFor(name[0])

                                        if text_instances:

                                            #split the email to get the first name
                                            first_name = name[1]
                                            first_name = first_name.split("@")
                                            first_name = first_name[0]
                                            first_name = first_name.split(".")
                                            first_name = first_name[0].capitalize()

                                            # send the email
                                            ret = send_email(email_sender,
                                                password_sender,
                                                name[1],
                                                'Timesheet signed',
                                                'Hello ' + first_name +',\nPlease find attached the Timesheet signed. Can you upload it in IDM?\n\nThank you very much. \nRegards, \nLaurence',
                                                output_path + '\\' + file)

                                            if ret == True:
                                                txt_show.insert(tk.END, "Email sent to: " + name[0] + " with email " + name[1] + "\n", 'OK')
                                            else:
                                                txt_show.insert(tk.END, "Email was not sent. Verify your credentials! " + "\n", 'error')
                                # close the file with the emails
                                emails_file.close()

                            # close the pdf
                            doc.close()

def clear_textarea():
    txt_show.delete('1.0', tk.END)

window = tk.Tk()
window.eval('tk::PlaceWindow . center')
window.geometry("700x310")
window.title("My signature app")
window.rowconfigure(0, minsize=200, weight=1)
window.columnconfigure(1, minsize=200, weight=1)

txt_show = tk.Text(window)
txt_show.tag_config('error', background="white", foreground="red")
txt_show.tag_config('warning', background="yellow", foreground="black")
txt_show.tag_config('OK', background="white", foreground="green")
fr_buttons = tk.Frame(window, relief=tk.RAISED, bd=2)


folderPath = tk.StringVar()
folder_label = tk.Label(fr_buttons, text="Select the folder")
folder_label.grid(row=0, column=0, sticky="ew")
e = tk.Entry(fr_buttons, textvariable=folderPath)
e.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
btn_find_folder = ttk.Button(fr_buttons, text="Browse",command=getFolderPath)
btn_find_folder.grid(row=1,column=1, sticky="ew")

signaturePath = tk.StringVar()
signature_label = tk.Label(fr_buttons ,text="Select the signature")
signature_label.grid(row=2,column = 0)
f = tk.Entry(fr_buttons,textvariable=signaturePath)
f.grid(row=3,column=0, sticky="ew", padx=5, pady=5)
btn_find_signature = ttk.Button(fr_buttons, text="Browse",command=getSignaturePath)
btn_find_signature.grid(row=3,column=1)

emailsPath = tk.StringVar()
email_label = tk.Label(fr_buttons ,text="Select the file with emails")
email_label.grid(row=7,column = 0)
f = tk.Entry(fr_buttons,textvariable=emailsPath)
f.grid(row=8,column=0, sticky="ew", padx=5, pady=5)
btn_find_emails = ttk.Button(fr_buttons, text="Browse",command=getEmailsPath)
btn_find_emails.grid(row=8,column=1)

btn_start = tk.Button(fr_buttons, text="Sign the documents", command = add_signatures)
btn_send = tk.Button(fr_buttons, text="Send the documents", command=send_emails_action)
btn_clear = tk.Button(fr_buttons, text="Clean textbox", command=clear_textarea)
btn_quit = tk.Button(fr_buttons, text="Exit", command=window.quit)


btn_start.grid(row=4, column=0, sticky="ew", padx=5, pady=5)
btn_send.grid(row=9, column=0, sticky="ew", padx=5)

tk.Label(fr_buttons, width=3, height=2).grid(column=0, row=10)

btn_clear.grid(row=11, column=0, sticky="ew", padx=5)
btn_quit.grid(row=12, column=0, sticky="ew", padx=5)

fr_buttons.grid(row=0, column=0, sticky="ns")
txt_show.grid(row=0, column=1, sticky="nsew")

window.mainloop()